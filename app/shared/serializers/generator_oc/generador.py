
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from openpyxl.drawing.image import Image


def _normalize_proveedor_name(proveedor: str) -> str:
    """
    Normaliza el nombre del proveedor para que sea seguro en URLs:
    - Elimina espacios al inicio y al final
    - Reemplaza espacios múltiples internos por un solo guión bajo
    - Conserva TODAS las palabras (S.A.C, E.I.R.L, etc.)
    """
    # Normalizar espacios: eliminar espacios al inicio/final
    proveedor = proveedor.strip()

    # Reemplazar espacios múltiples internos por un solo espacio y luego por guión bajo
    proveedor = ' '.join(proveedor.split())

    # Reemplazar todos los espacios por guiones bajos
    return proveedor.replace(' ', '_')


class Generador:
    def __init__(self, num_orden, oc, proveedor, igv, output_folder=None,consorcio=False): 
        # Si no se especifica output_folder, usar la carpeta data dentro del mismo directorio
        if output_folder is None:
            # Obtener el directorio actual del archivo generador.py
            current_dir = os.path.dirname(os.path.abspath(__file__))
            output_folder = os.path.join(current_dir, "data")
        
        # Crear la carpeta de salida si no existe
        os.makedirs(output_folder, exist_ok=True)
        
        self.cabecera = {
            "B11": "Señores :",
            "B12": "Atencion:",
            "B13": "Telefono:",
            "B14": "Correo :",
            "B15": "Direccion :",
            "F11": "Fecha :",
            "F12": "Moneda :",
            "F13": "Entrega :",
            "F14": "Pago :"
        }
        self.output_folder = output_folder
        self.orden = oc
        self.igv=igv
        self.moneda="SOLES"
        self.consorcio=consorcio
        # Configuración de la imagen
        current_dir = os.path.dirname(os.path.abspath(__file__))

        if self.consorcio:
            self.image_path = os.path.join(current_dir, "img", "consorcioLogo.png")
        else:
            self.image_path = os.path.join(current_dir, "img", "principal.png")
        
        
        self.image_width = 300  # Ancho en píxeles
        self.image_height = 150  # Alto en píxeles
        self.image_position = "D1"  # Posición en Excel (ej: "D1")



        # Normalizar el nombre del proveedor para evitar espacios en URLs
        proveedor_normalizado = _normalize_proveedor_name(proveedor)
        # Normalizar el número de orden también (eliminar espacios múltiples)
        num_orden_normalizado = '_'.join(num_orden.strip().split())
        self.output_file = f"{num_orden_normalizado}_{proveedor_normalizado}.xlsx"
        self.subTitle=f"ORDEN DE COMPRA N° {num_orden}"
        self.wb = Workbook()
        self.ws = self.wb.active
        self.last_product_row = None

    def aplicar_estilo_fondo(self):
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in self.ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
            for cell in row:
                cell.fill = white_fill

    def configurar_ancho_columna(self, col_index, width):
        col_letter = get_column_letter(col_index)
        self.ws.column_dimensions[col_letter].width = width

    def agregar_header(self):
        for cell, value in self.cabecera.items():
            self.ws[cell] = value
            self.ws[cell].font = Font(bold=True)

    def agregar_cont_header(self):
        datos = self.orden[0]  # Usar el primer registro para rellenar los valores de la cabecera
        self.ws["C11"] = datos.get("PROVEEDOR", "")
        self.ws["C12"] = datos.get("PERSONAL", "")
        self.ws["C13"] = datos.get("CELULAR", "")
        self.ws["C13"].alignment = Alignment(horizontal="left")
        self.ws["C14"] = datos.get("CORREO", "")
        self.ws["C15"] = datos.get("DIRECCION", "")
        self.ws["G11"] = datos.get("FECHA", "")
        self.ws["G11"].alignment = Alignment(horizontal="left")
        self.ws["G12"] = datos.get("MONEDA", "")
        self.moneda=datos.get("MONEDA", "SOLES")
        self.ws["G13"] = "INMEDIATO"
        self.ws["G14"] = datos.get("PAGO", "")
        self.ws["D8"] = self.subTitle
        self.ws["D8"].font = Font(bold=True)  
        self.ws["D8"].alignment = Alignment(horizontal="center")

    def agregar_productos(self):
        encabezados = ["CANT", "UND", "PRODUCTO", "MARCA", "MODELO", "P.UNIT", "P.TOTAL"]
        start_row = 18
        start_col = 2

        double_side = Side(border_style="double", color="000000")
        center_alignment = Alignment(horizontal='center', vertical='center')  # Alineación centrada para todo
        left_alignment = Alignment(horizontal='left', vertical='center')      # Alineación izquierda para producto
        #right_alignment = Alignment(horizontal='right', vertical='center')

        # Configurar encabezados
        for col_offset, encabezado in enumerate(encabezados):
            cell = self.ws.cell(row=start_row, column=start_col + col_offset)
            cell.value = encabezado
            cell.font = Font(bold=True)
            cell.border = Border(top=double_side, bottom=double_side)
            cell.alignment = left_alignment
            if col_offset == 2:  # Columna PRODUCTO
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment

        # Fijar altura de fila para encabezados
        self.ws.row_dimensions[start_row].height = 18

       # Determinar formato de contabilidad según moneda
        if self.moneda.upper() == "DOLARES":
            currency_format = '"$"#,##0.00'  # Formato para dólares
        else:
            # Formato contabilidad para soles: símbolo izquierda, número derecha
            currency_format = '"S/."#,##0.00;[Red]"S/."#,##0.00'

        # Procesar productos
        for idx, producto in enumerate(self.orden, start=1):
            row = start_row + idx
            
            # Fijar altura para cada fila de producto
            self.ws.row_dimensions[row].height = 35
            
            # Crear celdas con valores y alineaciones específicas
            cant_cell = self.ws.cell(row=row, column=2)
            cant_cell.value = producto.get("CANT", "")
            cant_cell.alignment = center_alignment
            
            und_cell = self.ws.cell(row=row, column=3)
            und_cell.value = producto.get("UMED", "")
            und_cell.alignment = center_alignment
            
            # Producto con alineación izquierda
            producto_cell = self.ws.cell(row=row, column=4)
            producto_cell.value = producto.get("PRODUCTO", "")
            producto_cell.alignment = left_alignment
            
            marca_cell = self.ws.cell(row=row, column=5)
            marca_cell.value = producto.get("MARCA", "")
            marca_cell.alignment = center_alignment

            modelo_cell = self.ws.cell(row=row, column=6)
            modelo_cell.value = producto.get("MODELO", "")
            modelo_cell.alignment = center_alignment

            # P.UNIT con alineación centrada
            punit_cell = self.ws.cell(row=row, column=7)
            punit_cell.value = producto.get("P.UNIT", "")
            punit_cell.alignment = center_alignment
            
            # P.TOTAL con formato de moneda
            ptotal_cell = self.ws.cell(row=row, column=8)
            ptotal_cell.value = f"=B{row}*G{row}"
            ptotal_cell.number_format = currency_format
            ptotal_cell.alignment = center_alignment
            


        self.last_product_row = start_row + len(self.orden)

       # Configurar borde inferior
        if self.last_product_row > start_row:
            self.ws.row_dimensions[self.last_product_row].height = 25
            for col_offset in range(len(encabezados)):
                cell = self.ws.cell(row=self.last_product_row, column=start_col + col_offset)
                cell.border = Border(bottom=double_side)
                if col_offset == 0:  # Columna CANT
                    cell.alignment = center_alignment
                elif col_offset == 1:  # Columna UND
                    cell.alignment = center_alignment
                elif col_offset == 2:  # Columna PRODUCTO
                    cell.alignment = left_alignment
                elif col_offset == 3:  # Columna MARCA
                    cell.alignment = center_alignment
                elif col_offset == 4:  # Columna MODELO
                    cell.alignment = center_alignment
                elif col_offset == 5:  # Columna P.UNIT
                    cell.alignment = center_alignment
                elif col_offset == 6:  # Columna P.TOTAL
                    cell.alignment = center_alignment
                else:
                    cell.alignment = center_alignment

    def agregar_total(self):
        if not self.last_product_row:
            self.last_product_row = 18

        # Determinar formato de contabilidad según moneda
        if self.moneda.upper() == "DOLARES":
            currency_format = '"$"#,##0.00'  # Formato para dólares
        else:
            currency_format = '"S/."#,##0.00;[Red]"S/."#,##0.00'
            
        total_row = self.last_product_row + 1
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Crear celdas de totales
        total_formula_cell = self.ws.cell(row=total_row, column=8)
        total_formula_cell.value = f"=SUM(H18:H{self.last_product_row})"
        total_formula_cell.font = Font(bold=True)
        total_formula_cell.alignment = center_alignment
        total_formula_cell.number_format = currency_format
        
        if self.igv == "SIN IGV":
            cell_igv = self.ws.cell(row=total_row+1, column=8)
            cell_igv.value = f"=H{total_row}*0.18"
            cell_igv.font = Font(bold=True)
            cell_igv.alignment = center_alignment
            cell_igv.number_format = currency_format
            
            total_igv = self.ws.cell(row=total_row+2, column=8)
            total_igv.value= f"=SUM(H{total_row}:H{total_row+1})"
            total_igv.font = Font(bold=True)
            total_igv.alignment = center_alignment
            total_igv.number_format = currency_format
            
            # Crear descripciones
            self.ws.cell(row=total_row, column=7).value = "P.UNITARIO:"
            self.ws.cell(row=total_row, column=7).font = Font(bold=True)
            self.ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
            
            self.ws.cell(row=total_row+1, column=7).value = "IGV:"
            self.ws.cell(row=total_row+1, column=7).font = Font(bold=True)
            self.ws.cell(row=total_row+1, column=7).alignment = Alignment(horizontal="right")
            
            self.ws.cell(row=total_row+2, column=7).value = "TOTAL:"
            self.ws.cell(row=total_row+2, column=7).font = Font(bold=True)
            self.ws.cell(row=total_row+2, column=7).alignment = Alignment(horizontal="right")
            self.ws.cell(row=total_row+2, column=7).border = Border(top=Side(style='thin'))
        else:
            self.ws.cell(row=total_row, column=7).value = "TOTAL:"
            self.ws.cell(row=total_row, column=7).font = Font(bold=True)
            self.ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")   

        
    def agregar_footer(self):
        if not self.last_product_row:
            self.last_product_row = 18

        info_start_row = self.last_product_row + 4  


        info_lines_corpelima = [
            "FACTURAR:",
            "CORPELIMA S.A.C",
            "RUC: 20601460913",
            "CALLE  55 MZ WW2 LT 13   URB. LA FLORESTA DE PRO   LOS OLIVOS  LIMA",
            "",
            "BCP:   CTA. CTE  DOLARES  Nro.    ",
            "BCP:   CTA. CTE SOLES     Nro.   ",
            "",
            "CEL:  926502167",
            "ventas@corpelima.com"
        ]

        info_lines_consorcio=[
            "FACTURAR:",
            "CONSORCIO ELECTRO MINERO S.A.C",
            "RUC: 20565501942",
            "CALLE  55 MZ WW2 LT 13   URB. LA FLORESTA DE PRO   LOS OLIVOS  LIMA",
            "",
            "BCP:   CTA. CTE  DOLARES  Nro.    ",
            "BCP:   CTA. CTE SOLES     Nro.    ",
            "",
            "CEL:  926502167",
            "ventas.consorcioelectrominero@gmail.com"
        ]

        if self.consorcio:
            info_lines = info_lines_consorcio
        else:
            info_lines = info_lines_corpelima


        for i, line in enumerate(info_lines):
            row = info_start_row + i
            self.ws.cell(row=row, column=4).value = line
            self.ws.cell(row=row, column=4).font = Font(bold=True)
            #self.ws.cell(row=row, column=4).font.size = 10

    def agregar_imagen(self):
        """
        Agrega la imagen al Excel con el tamaño y posición especificados
        """
        try:
            img = Image(self.image_path)
             
            # Configurar el tamaño de la imagen si se especifica
            if self.image_width and self.image_height:
                # Convertir píxeles a puntos (1 píxel ≈ 0.75 puntos)
                width_pt = int(self.image_width * 1.80)
                height_pt = int(self.image_height * 0.75)
                img.width = width_pt
                img.height = height_pt
                ##print(f"Imagen configurada con tamaño: {self.image_width}x{self.image_height} píxeles ({width_pt:.1f}x{height_pt:.1f} puntos)")
            elif self.image_width:
                # Solo ancho especificado, mantener proporción
                width_pt = int(self.image_width * 0.75)
                img.width = width_pt
                #print(f"Imagen configurada con ancho: {self.image_width} píxeles ({width_pt:.1f} puntos)")
            elif self.image_height:
                # Solo alto especificado, mantener proporción
                height_pt = int(self.image_height * 0.75)
                img.height = height_pt
                #print(f"Imagen configurada con alto: {self.image_height} píxeles ({height_pt:.1f} puntos)")
            else:
                # Tamaño por defecto (pequeño)
                img.width = 150  # 200 píxeles
                img.height = 75  # 100 píxeles
                #print("Imagen configurada con tamaño por defecto: 200x100 píxeles")
            
            # Agregar la imagen en la posición especificada
            self.ws.add_image(img, self.image_position)
            #print(f"Imagen agregada en posición: {self.image_position}")
            
        except Exception as e:
            print(f"Advertencia: No se pudo cargar la imagen {self.image_path}: {e}")

    def guardar_archivo(self):
        full_path = os.path.join(self.output_folder, self.output_file)
        self.wb.save(full_path)
        print(f"Archivo '{self.output_file}' creado con éxito en: {full_path}")

    def generar_excel(self):
        self.aplicar_estilo_fondo()
        self.configurar_ancho_columna(3, 11)
        self.configurar_ancho_columna(4, 75)
        self.configurar_ancho_columna(5, 15)
        self.configurar_ancho_columna(6, 12)
        self.configurar_ancho_columna(7, 12)
        self.configurar_ancho_columna(8, 12)
        self.agregar_header()
        self.agregar_cont_header()
        self.agregar_productos()
        self.agregar_total()  
        self.agregar_footer()
        self.agregar_imagen()
        self.guardar_archivo()
